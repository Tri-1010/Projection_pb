"""
Test script for enhanced lifecycle export
"""
import sys
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime

# Add project root to path
project_root = Path(".").resolve()
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from src.rollrate.lifecycle_export_enhanced import export_lifecycle_with_config_info

# Create sample data
print("📦 Creating sample data...")

# Sample lifecycle data
dates = pd.date_range('2023-01-01', periods=12, freq='MS')
mobs = list(range(1, 14))
products = ['C', 'S', 'T']

data = []
for product in products:
    for vintage in dates:
        for mob in mobs:
            data.append({
                'PRODUCT_TYPE': product,
                'VINTAGE_DATE': vintage,
                'MOB': mob,
                'DEL30_PCT': np.random.uniform(0.01, 0.15),
                'DEL60_PCT': np.random.uniform(0.005, 0.10),
                'DEL90_PCT': np.random.uniform(0.002, 0.08),
                'IS_FORECAST': 1 if mob > 6 else 0,
            })

df_del_prod = pd.DataFrame(data)

# Sample raw data
raw_data = []
for i in range(1000):
    raw_data.append({
        'AGREEMENT_ID': f'LOAN_{i:06d}',
        'CUTOFF_DATE': 202401,
        'DISBURSAL_DATE': pd.Timestamp('2023-01-01') + pd.Timedelta(days=i % 365),
        'DISBURSAL_AMOUNT': np.random.uniform(10000, 100000),
        'PRINCIPLE_OUTSTANDING': np.random.uniform(5000, 90000),
        'MOB': np.random.randint(1, 13),
        'PRODUCT_TYPE': np.random.choice(products),
        'RISK_SCORE': np.random.choice(['500-', '550-', '600-', '650+']),
        'STATE_MODEL': np.random.choice(['DPD0', 'DPD30+', 'DPD60+', 'DPD90+']),
    })

df_raw = pd.DataFrame(raw_data)

# Sample actual_info
actual_info = {}
for product in products:
    for vintage in dates:
        actual_info[(product, vintage)] = 6  # Actual data up to MOB 6

# Config params
config_params = {
    'DATA_PATH': 'C:/Users/User/Projection_PB/Projection_pb/POS_Parquet_YYYYMM',
    'MAX_MOB': 13,
    'TARGET_MOBS': [12],
    'SEGMENT_COLS': ['PRODUCT_TYPE', 'RISK_SCORE', 'GENDER', 'LA_GROUP', 'SALE_CHANNEL'],
    'MIN_OBS': 100,
    'MIN_EAD': 100,
    'WEIGHT_METHOD': 'exp',
    'ROLL_WINDOW': 20,
    'DECAY_LAMBDA': 0.97,
}

# Test export
print("📊 Testing export...")
output_file = "test_Lifecycle_All_Products.xlsx"

try:
    export_lifecycle_with_config_info(
        df_del_prod=df_del_prod,
        actual_info=actual_info,
        df_raw=df_raw,
        config_params=config_params,
        filename=output_file
    )
    
    print(f"\n✅ Test successful! File created: {output_file}")
    print(f"   File size: {Path(output_file).stat().st_size / 1024:.1f} KB")
    
    # Verify sheets
    import openpyxl
    wb = openpyxl.load_workbook(output_file, read_only=True)
    print(f"\n📋 Sheets in file:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"   {i}. {sheet_name}")
    
    if 'Config_Info' in wb.sheetnames:
        print("\n✅ Config_Info sheet found!")
    else:
        print("\n❌ Config_Info sheet NOT found!")
    
except Exception as e:
    print(f"\n❌ Test failed: {e}")
    import traceback
    traceback.print_exc()
